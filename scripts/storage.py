#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 存储管理模块
负责 jobs.xlsx 的创建、读取、写入、去重。

新增特性：
  - 「投递状态」列（待投递 / 已投递 / 已过期 / 不合适）
  - 投递方式结构化格式化（电话/邮箱/网上报名/现场报名）
  - 统一使用 logger 模块
  - get_all_jobs() 返回完整数据供 query.py 使用
  - update_field() 支持按行号更新单个字段（用于 rescan）
"""
import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse, parse_qs

from scripts.logger import get_logger

log = get_logger("storage")

# ── 路径 ──────────────────────────────────────────────────────────
EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "jobs.xlsx")

# ── 表头（顺序即列顺序，新增「投递状态」） ────────────────────────
HEADERS = [
    "收藏时间",
    "岗位名称",
    "招聘单位",
    "所在地区",
    "招录人数",
    "学历要求",
    "专业要求",
    "报名截止日期",
    "投递方式",
    "考试类型",
    "其他要求",
    "备注",
    "投递状态",       # 新增列
    "原始链接",
]

# 投递状态可选值
DELIVERY_STATUS_OPTIONS = ["待投递", "已投递", "已过期", "不合适"]
DELIVERY_STATUS_DEFAULT  = "待投递"

# 列宽配置
_COL_WIDTHS = {
    "收藏时间":    18,
    "岗位名称":    20,
    "招聘单位":    28,
    "所在地区":    14,
    "招录人数":     8,
    "学历要求":    15,
    "专业要求":    22,
    "报名截止日期": 14,
    "投递方式":    32,
    "考试类型":    12,
    "其他要求":    22,
    "备注":        22,
    "投递状态":    12,
    "原始链接":    45,
}

# 投递状态单元格颜色
_STATUS_COLORS = {
    "待投递": "FFF2CC",   # 淡黄
    "已投递": "D9EAD3",   # 淡绿
    "已过期": "F4CCCC",   # 淡红
    "不合适": "EFEFEF",   # 灰
}


# ── 投递方式格式化 ────────────────────────────────────────────────

def normalize_delivery(raw: str) -> str:
    """
    标准化投递方式：统一格式为 [类型]：[详情]，多项用换行分隔。

    示例：
      "联系电话：010-12345678" → "电话：010-12345678"
      "hr@example.com"        → "邮箱：hr@example.com"
      "https://xxx.com/apply" → "网上报名：https://xxx.com/apply"
    """
    if not raw or str(raw).strip() in ("", "暂无", "nan"):
        return "暂无"

    text = str(raw).strip()

    # 拆分多个条目（逗号/分号/换行分隔）
    parts = re.split(r"[,，;\n]+", text)
    normalized = []

    for part in parts:
        part = part.strip()
        if not part:
            continue

        # ── 电话 ──
        part = re.sub(r"^(?:联系电话|电话|TEL|手机|Mobile|联系方式)[：:\s]+",
                      "", part, flags=re.IGNORECASE)
        phone_re = re.compile(r"^[\d\-\(\)\s+]+$")
        if phone_re.match(part):
            normalized.append("电话：" + part.strip())
            continue

        # ── 邮箱 ──
        email_re = re.compile(r"[\w.+-]+@[\w-]+\.[a-z]{2,}", re.IGNORECASE)
        if email_re.search(part):
            email = email_re.search(part).group()
            normalized.append("邮箱：" + email)
            continue

        # ── 网上报名（含 URL）──
        if re.search(r"https?://", part):
            url_match = re.search(r"https?://\S+", part)
            url = url_match.group() if url_match else part
            normalized.append("网上报名：" + url)
            continue

        # ── 现场报名关键词 ──
        if re.search(r"现场|地址|地点", part):
            normalized.append("现场报名：" + part)
            continue

        # ── 其他：保留原文 ──
        normalized.append(part)

    return "，".join(normalized) if normalized else "暂无"


# ── Excel 工作簿管理 ──────────────────────────────────────────────

def _create_workbook() -> Workbook:
    """创建新的 Excel 工作簿，写入表头并设置样式"""
    wb = Workbook()
    ws = wb.active
    ws.title = "岗位收藏"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"

    for col, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = _COL_WIDTHS.get(header, 15)

    ws.row_dimensions[1].height = 25
    return wb


def _normalize_url_for_dedup(url: str) -> str:
    """规范化 URL 用于去重（去掉追踪参数，保留业务 ID）"""
    if not url:
        return ""
    parsed = urlparse(url)

    hash_query_str = ""
    if parsed.fragment and "?" in parsed.fragment:
        _, hash_query_str = parsed.fragment.split("?", 1)

    all_params = parse_qs(parsed.query)
    if hash_query_str:
        all_params.update(parse_qs(hash_query_str))

    stable = ["id", "jobId", "article", "job", "position", "page", "articleId", "job_id"]
    filtered = {k: v for k, v in all_params.items() if k in stable}

    if filtered:
        query = "&".join(f"{k}={v[0]}" for k, v in filtered.items())
        return f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{query}"
    return f"{parsed.scheme}://{parsed.netloc}{parsed.path}"


def _is_duplicate(ws, url: str) -> bool:
    """检查链接是否已存在"""
    url_col = HEADERS.index("原始链接") + 1  # 0-indexed position + 1 for 1-indexed column
    normalized = _normalize_url_for_dedup(url)
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 安全检查：确保行数据足够长
        if url_col - 1 >= len(row):
            continue
        existing = row[url_col - 1]
        if existing and _normalize_url_for_dedup(str(existing)) == normalized:
            return True
    return False


def _set_status_cell_color(cell, status: str):
    """根据投递状态设置单元格背景色"""
    color = _STATUS_COLORS.get(status, "FFFFFF")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


# ── 公开 API ──────────────────────────────────────────────────────

def save_job(job_info: dict) -> dict:
    """
    将岗位信息追加写入 Excel。

    Returns:
        {"success": bool, "message": str, "row": int}
    """
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        log.info(f"创建新的 Excel 文件: {EXCEL_FILE}")
        wb = _create_workbook()
        ws = wb.active

    # ── 去重检查 ──
    url = job_info.get("原始链接", "")
    if url and _is_duplicate(ws, url):
        return {"success": False, "message": "该链接已存在，跳过重复收藏", "row": -1}

    # ── 格式化投递方式 ──
    if "投递方式" in job_info:
        job_info["投递方式"] = normalize_delivery(job_info["投递方式"])

    # ── 构建行数据 ──
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    row_data = [now]
    for header in HEADERS[1:]:
        if header == "投递状态":
            row_data.append(DELIVERY_STATUS_DEFAULT)
        else:
            row_data.append(job_info.get(header, ""))

    # ── 写入 ──
    next_row = ws.max_row + 1
    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 投递状态列染色
        header = HEADERS[col - 1]
        if header == "投递状态":
            _set_status_cell_color(cell, str(value))
        elif next_row % 2 == 0:
            cell.fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")

    ws.row_dimensions[next_row].height = 20

    # 保存时处理文件被占用等异常
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        log.fail(f"保存失败：jobs.xlsx 正被其他程序占用，请关闭后再试")
        return {"success": False, "message": "jobs.xlsx 正被其他程序占用", "row": -1}
    except Exception as e:
        log.fail(f"保存失败：{e}")
        return {"success": False, "message": f"保存失败：{e}", "row": -1}

    return {
        "success": True,
        "message": f"已成功保存到 jobs.xlsx（第 {next_row - 1} 条）",
        "row": next_row - 1,
    }


def get_stats() -> dict:
    """获取收藏统计信息"""
    if not os.path.exists(EXCEL_FILE):
        return {"total": 0, "file": EXCEL_FILE}

    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb.active
    total = max(0, ws.max_row - 1)
    wb.close()

    return {"total": total, "file": EXCEL_FILE}


def get_all_jobs() -> list[dict]:
    """
    读取 jobs.xlsx 中所有记录，返回字典列表。
    每条记录额外附加 "_row" 键（Excel 行号，从 2 开始）。
    """
    if not os.path.exists(EXCEL_FILE):
        return []

    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb.active

    jobs = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        job = {"_row": row_idx}
        for col, header in enumerate(HEADERS):
            val = row[col] if col < len(row) else ""
            job[header] = val if val is not None else ""
        jobs.append(job)

    wb.close()
    return jobs


def update_field(row_idx: int, field: str, value) -> bool:
    """
    更新指定行的单个字段。

    Args:
        row_idx: Excel 行号（从 2 开始，1 为表头）
        field:   字段名（必须在 HEADERS 中）
        value:   新值

    Returns:
        True 表示成功，False 表示字段不存在
    """
    if field not in HEADERS:
        log.warn(f"字段 '{field}' 不在 HEADERS 中，跳过更新")
        return False

    if not os.path.exists(EXCEL_FILE):
        log.error("jobs.xlsx 不存在，无法更新")
        return False

    col_idx = HEADERS.index(field) + 1
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    cell = ws.cell(row=row_idx, column=col_idx, value=value)
    if field == "投递状态":
        _set_status_cell_color(cell, str(value))

    # 保存时处理文件被占用等异常
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        log.fail(f"更新失败：jobs.xlsx 正被其他程序占用，请关闭后再试")
        return False
    except Exception as e:
        log.fail(f"更新失败：{e}")
        return False
    return True


def update_delivery_status(row_idx: int, status: str) -> bool:
    """便捷方法：更新投递状态"""
    if status not in DELIVERY_STATUS_OPTIONS:
        log.warn(f"无效的投递状态: {status}，可选值: {DELIVERY_STATUS_OPTIONS}")
        return False
    return update_field(row_idx, "投递状态", status)


# ── 模块自测 ──────────────────────────────────────────────────────
if __name__ == "__main__":
    test_job = {
        "岗位名称":    "综合管理岗",
        "招聘单位":    "XX市人力资源和社会保障局",
        "所在地区":    "广东省广州市",
        "招录人数":    "2",
        "学历要求":    "本科及以上",
        "专业要求":    "行政管理、公共管理类",
        "报名截止日期": "2026-05-15",
        "投递方式":    "联系电话：010-12345678, hr@test.com, https://apply.example.com",
        "考试类型":    "省考",
        "其他要求":    "年龄35岁以下",
        "备注":       "需要笔试+面试",
        "原始链接":    "https://www.gongkaoleida.com/test/12345",
    }
    result = save_job(test_job)
    log.ok(result["message"])

    stats = get_stats()
    log.info(f"当前共 {stats['total']} 条记录，文件: {stats['file']}")

    jobs = get_all_jobs()
    if jobs:
        log.info(f"第一条记录: {jobs[0].get('岗位名称')} | 投递状态: {jobs[0].get('投递状态')}")
