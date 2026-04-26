#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
历史数据回归脚本
重新抓取 jobs.xlsx 中的历史记录，用 LLM 重新解析，更新字段
"""
import os
import sys
import re
import json
import time
import subprocess
from datetime import datetime
from dotenv import load_dotenv

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env"))

from openpyxl import load_workbook
import pandas as pd

from scripts.parser import parse_job_info

EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "jobs.xlsx")


def normalize_人数(value) -> str:
    """标准化招录人数：若干→0"""
    if pd.isna(value):
        return 0
    val = str(value).strip()
    if val in ("若干", "若干人", "不限", "若干名"):
        return 0
    try:
        return int(re.search(r"\d+", val).group())
    except (AttributeError, ValueError):
        return 0


def normalize_投递方式(value) -> str:
    """标准化投递方式格式"""
    if pd.isna(value) or not str(value).strip() or str(value).strip() == "暂无":
        return "暂无"

    text = str(value).strip()

    # 去掉"联系电话："前缀，统一格式
    text = re.sub(r"^联系电话[：:]\s*", "", text)

    # 去掉"电话："、"TEL："等前缀
    text = re.sub(r"^(电话|TEL|手机|Mobile)[：:]\s*", "", text, flags=re.IGNORECASE)

    # 清理多余空格
    text = re.sub(r"\s+", " ", text).strip()

    # 如果是纯电话号码，加上"电话："前缀
    phone_pattern = r"^[\d\-\(\)\s,，]+$"
    if re.match(phone_pattern, text):
        return "电话：" + text

    return text


def is_old_format_remark(remark: str) -> bool:
    """判断备注是否为旧格式（职位介绍原文）"""
    if pd.isna(remark) or not str(remark).strip():
        return True

    text = str(remark).strip()

    # 新格式应该包含亮点关键词，而不是职位描述
    old_format_indicators = [
        "负责",
        "职位介绍",
        "智能监控体系建设",
        "负责公司信息化",
        "负责智能化矿山",
        "负责书院水电",
        "开展",
        "推动",
        "搭建",
    ]

    for indicator in old_format_indicators:
        if text.startswith(indicator):
            return True

    # 如果备注以编号列表开头（如 "1.负责..."），判定为旧格式
    if re.match(r"^\d+[.、]", text):
        return True

    return False


def rescan_single_row(row_idx: int, url: str) -> dict | None:
    """重新抓取并解析单个记录"""
    print(f"\n[{row_idx}] 正在处理: {url}")

    try:
        # 1. 抓取页面（使用 subprocess 调用 scraper.py 避免 asyncio 冲突）
        script_path = os.path.join(os.path.dirname(__file__), "scraper.py")
        result = subprocess.run(
            [sys.executable, script_path, url],
            capture_output=True,
            text=True,
            timeout=60,
        )

        if result.returncode != 0:
            print(f"[{row_idx}] 抓取失败: {result.stderr}")
            return None

        # scraper.py 输出格式：标题、URL、文本预览
        # 我们需要获取完整文本，用 --full 参数或其他方式
        # 简化处理：直接读取 scraper.py 输出的 JSON

        # 实际上 scraper.py 没有 JSON 输出，我们需要修改它或者手动解析
        # 这里改为直接调用 async 函数
        import asyncio
        from scripts.scraper import scrape_page

        page_data = asyncio.run(scrape_page(url))

        # 2. LLM 解析
        job_info = parse_job_info(page_data)

        # 3. 标准化字段
        招录人数 = normalize_人数(job_info.get("招录人数", 0))
        投递方式 = normalize_投递方式(job_info.get("投递方式", "暂无"))
        备注 = job_info.get("备注", "暂无")

        return {
            "招录人数": 招录人数,
            "投递方式": 投递方式,
            "备注": 备注,
        }

    except Exception as e:
        print(f"[{row_idx}] 处理失败: {e}")
        import traceback
        traceback.print_exc()
        return None


def rescan_all():
    """主函数：重新处理所有历史记录"""
    print("=" * 60)
    print("历史数据回归脚本")
    print("=" * 60)

    # 读取 Excel
    df = pd.read_excel(EXCEL_FILE)
    total = len(df)

    if total == 0:
        print("没有找到历史数据")
        return

    print(f"共找到 {total} 条历史记录")

    # 统计需要更新的记录
    need_update = 0
    for idx, row in df.iterrows():
        remark = row.get("备注", "")
        投递方式 = row.get("投递方式", "")
        招录人数 = row.get("招录人数", "")

        if is_old_format_remark(remark) or str(投递方式) == "nan" or str(招录人数) == "若干":
            need_update += 1
            print(f"\n[记录 {idx + 1}] 需要更新:")
            print(f"  备注: {str(remark)[:80]}...")
            print(f"  投递方式: {投递方式}")
            print(f"  招录人数: {招录人数}")

    print(f"\n需要更新的记录: {need_update}/{total}")
    print("\n开始重新抓取和解析...\n")

    # 逐条处理（避免并发导致 session 问题）
    results = []
    for idx, row in df.iterrows():
        url = row.get("原始链接", "")

        if not url or str(url) == "nan":
            print(f"[跳过] 第 {idx + 1} 条: 无有效链接")
            results.append(None)
            continue

        result = rescan_single_row(idx + 1, url)
        results.append(result)

        # 避免请求过快
        if idx < total - 1:
            time.sleep(3)

    # 写入更新
    print("\n" + "=" * 60)
    print("更新 Excel 文件...")
    print("=" * 60)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    updated_count = 0
    for idx, (result, row) in enumerate(zip(results, df.itertuples()), start=0):
        excel_row = idx + 2  # Excel 行号（跳过表头）

        if result is None:
            print(f"[{idx + 1}] 跳过（无结果）")
            continue

        # 更新招录人数
        ws.cell(row=excel_row, column=5, value=result["招录人数"])

        # 更新投递方式
        ws.cell(row=excel_row, column=9, value=result["投递方式"])

        # 更新备注
        ws.cell(row=excel_row, column=12, value=result["备注"])

        updated_count += 1
        print(f"[{idx + 1}] 已更新: 招录人数={result['招录人数']}, 投递方式={result['投递方式'][:30]}...")

    wb.save(EXCEL_FILE)
    print(f"\n完成！已更新 {updated_count} 条记录")


def main():
    rescan_all()


if __name__ == "__main__":
    main()
